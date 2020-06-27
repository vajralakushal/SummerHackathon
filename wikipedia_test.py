import wikipediaapi
import openpyxl
import time

# Setting up all of the files/objects that need to be used
clean_file = open("clean_data.txt", 'r')
simple_wiki = wikipediaapi.Wikipedia('simple')
normal_wiki = wikipediaapi.Wikipedia('en')
workbook = openpyxl.load_workbook("data.xlsx")
worksheet = workbook['summaries']

#loop that puts everything in
start_time = time.time()
current_row = 2
current_col = 1
ID = 0
counter = 0
while counter != 100:
    cell = worksheet.cell(row = current_row, column = current_col)
    line = clean_file.readline().rstrip("\n")
    if not line:
        break
    simple_page = simple_wiki.page(line)
    normal_page = normal_wiki.page(line)
    if simple_page.exists() and normal_page.exists():
        cell.value = line
        current_col += 1
        cell = worksheet.cell(row = current_row, column = current_col)
        simple_summary = simple_page.summary[:]
        normal_summary = normal_page.summary[:]
        if "\n" in simple_summary:
            cell.value = simple_summary.replace("\n", "\a").split("\a")[0]
        else:
            cell.value = simple_summary
            
        current_col += 1
        cell = worksheet.cell(row = current_row, column = current_col)
        if "\n" in normal_summary:
            cell.value = normal_summary.replace("\n", "\a").split("\a")[0]
        else:
            cell.value = normal_summary
        
        current_col += 1
        cell = worksheet.cell(row = current_row, column = current_col)
        cell.value = ID
        ID += 1
        
        
        current_row += 1
        current_col -= 3
    workbook.save('data.xlsx')
    counter += 1
print("--- %s seconds ---" % (time.time() - start_time))