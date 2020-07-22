import openpyxl

workbook = openpyxl.load_workbook("data.xlsx")
worksheet = workbook['summaries']
current_row = 2
current_col = 2
counter = 0

def isEnglish(s):
    try:
        s.encode(encoding='utf-8').decode('ascii')
    except UnicodeDecodeError:
        return False
    else:
        return True

while counter != 100:
    simple_cell = str(worksheet.cell(row = current_row, column = current_col))
    normal_cell = str(worksheet.cell(row = current_row, column = current_col + 1))
    if simple_cell.isascii() and normal_cell.isascii():
        cell = worksheet.cell(row = current_row, column = current_col + 3)
        cell.value = 1
        print("yes " + str(current_row))
    else:
        cell = worksheet.cell(row = current_row, column = current_col + 3)
        cell.value = 0
    counter += 1
    current_row += 1




